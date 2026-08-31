from __future__ import annotations

import threading
from pathlib import Path

import openpyxl

APP_DIR = Path(__file__).resolve().parent
QUESTIONS_XLSX_PATH = APP_DIR.parent / "SCT Questions.xlsx"

_lock = threading.Lock()
_cache_mtime: float | None = None
_cache_questions: dict[tuple[str, int], str] = {}


def _load_questions() -> dict[tuple[str, int], str]:
    questions: dict[tuple[str, int], str] = {}
    workbook = openpyxl.load_workbook(QUESTIONS_XLSX_PATH, data_only=True, read_only=True)
    try:
        for age_group in workbook.sheetnames:
            sheet = workbook[age_group]
            for question_number, question_text in sheet.iter_rows(values_only=True):
                if question_number is None or question_text is None:
                    continue
                questions[(age_group, int(question_number))] = str(question_text).strip()
    finally:
        workbook.close()
    return questions


def get_question_text(age_group: str | None, question_number: int | None) -> str | None:
    """검사유형/문항번호로 SCT Questions.xlsx에서 질문 텍스트를 찾는다.

    파일 수정시각이 바뀌면 자동으로 다시 읽어, 서버 재시작 없이 엑셀 수정이 반영된다.
    """
    if not age_group or question_number is None:
        return None

    global _cache_mtime, _cache_questions
    try:
        mtime = QUESTIONS_XLSX_PATH.stat().st_mtime
    except OSError:
        return None

    with _lock:
        if mtime != _cache_mtime:
            _cache_questions = _load_questions()
            _cache_mtime = mtime
        return _cache_questions.get((age_group, question_number))
